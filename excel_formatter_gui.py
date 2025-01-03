import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
import threading
import time
from datetime import datetime, timedelta


class ExcelFormatterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel格式化工具")
        self.root.geometry("600x400")
        self.setup_ui()
        self.is_running = False

    def setup_ui(self):
        # 文件选择区域
        file_frame = ttk.LabelFrame(self.root, text="文件选择", padding="5")
        file_frame.pack(fill="x", padx=5, pady=5)

        self.file_path = tk.StringVar()
        self.file_entry = ttk.Entry(file_frame, textvariable=self.file_path, width=50)
        self.file_entry.pack(side="left", padx=5)

        self.browse_btn = ttk.Button(file_frame, text="浏览", command=self.browse_file)
        self.browse_btn.pack(side="left", padx=5)

        # 进度显示区域
        progress_frame = ttk.LabelFrame(self.root, text="处理进度", padding="5")
        progress_frame.pack(fill="x", padx=5, pady=5)

        # 总进度条
        ttk.Label(progress_frame, text="总体进度:").pack(fill="x")
        self.total_progress = ttk.Progressbar(
            progress_frame, length=400, mode="determinate"
        )
        self.total_progress.pack(fill="x", padx=5, pady=2)

        # 当前工作表进度条
        ttk.Label(progress_frame, text="当前工作表进度:").pack(fill="x")
        self.sheet_progress = ttk.Progressbar(
            progress_frame, length=400, mode="determinate"
        )
        self.sheet_progress.pack(fill="x", padx=5, pady=2)

        # 状态显示
        self.status_text = tk.Text(progress_frame, height=10, width=50)
        self.status_text.pack(fill="both", padx=5, pady=5)

        # 预计剩余时间
        self.time_label = ttk.Label(progress_frame, text="预计剩余时间: --:--")
        self.time_label.pack(fill="x", padx=5)

        # 控制按钮
        button_frame = ttk.Frame(self.root)
        button_frame.pack(fill="x", padx=5, pady=5)

        self.start_btn = ttk.Button(
            button_frame, text="开始处理", command=self.start_processing
        )
        self.start_btn.pack(side="left", padx=5)

        self.cancel_btn = ttk.Button(
            button_frame, text="取消", command=self.cancel_processing, state="disabled"
        )
        self.cancel_btn.pack(side="left", padx=5)

    def browse_file(self):
        filename = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.file_path.set(filename)

    def update_status(self, message):
        self.status_text.insert(
            "end", f"{datetime.now().strftime('%H:%M:%S')} - {message}\n"
        )
        self.status_text.see("end")

    def format_excel(self):
        try:
            filepath = self.file_path.get()
            wb = load_workbook(filepath)
            total_sheets = len(wb.sheetnames)

            start_time = time.time()

            for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
                if not self.is_running:
                    break

                ws = wb[sheet_name]
                self.update_status(f"开始处理工作表: {sheet_name}")

                # 更新总进度
                self.total_progress["value"] = (sheet_idx / total_sheets) * 100

                # 设置格式
                steps = [
                    ("设置首行格式", 20),
                    ("调整单元格对齐", 40),
                    ("添加框线", 60),
                    ("调整行高列宽", 80),
                    ("完成工作表处理", 100),
                ]

                # 冻结首行和筛选
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions

                # 设置首行格式
                first_row_font = Font(bold=True, color="000000")
                first_row_fill = PatternFill(
                    start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
                )

                # 边框样式
                thin_border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                # 获取已使用范围
                max_row = ws.max_row
                max_col = ws.max_column

                # 处理每个步骤
                for step_desc, progress in steps:
                    if not self.is_running:
                        break

                    self.update_status(step_desc)
                    self.sheet_progress["value"] = progress

                    # 更新预计剩余时间
                    elapsed_time = time.time() - start_time
                    sheets_remaining = total_sheets - sheet_idx
                    if sheet_idx > 0:
                        time_per_sheet = elapsed_time / sheet_idx
                        eta = time_per_sheet * sheets_remaining
                        eta_text = str(timedelta(seconds=int(eta)))
                        self.time_label["text"] = f"预计剩余时间: {eta_text}"

                    # 实际处理步骤
                    if "首行格式" in step_desc:
                        for cell in ws[1]:
                            cell.font = first_row_font
                            cell.fill = first_row_fill

                    elif "对齐" in step_desc:
                        for row in ws.iter_rows(
                            min_row=1, max_row=max_row, max_col=max_col
                        ):
                            for cell in row:
                                cell.alignment = Alignment(
                                    horizontal="center", vertical="center"
                                )

                    elif "框线" in step_desc:
                        for row in ws.iter_rows(
                            min_row=1, max_row=max_row, max_col=max_col
                        ):
                            for cell in row:
                                cell.border = thin_border

                    elif "行高列宽" in step_desc:
                        # 调整列宽
                        for col in range(1, max_col + 1):
                            max_length = 0
                            column = get_column_letter(col)
                            for row in range(1, max_row + 1):
                                cell = ws[f"{column}{row}"]
                                if cell.value:
                                    cell_length = sum(
                                        2 if ord(char) > 127 else 1
                                        for char in str(cell.value)
                                    )
                                    max_length = max(max_length, cell_length)
                            ws.column_dimensions[column].width = max_length + 2

                        # 调整行高
                        for row in range(1, max_row + 1):
                            ws.row_dimensions[row].height = 20

                    self.root.update()
                    time.sleep(0.1)  # 添加小延迟使进度更平滑

            if self.is_running:
                wb.save(filepath)
                self.update_status("处理完成！")
                messagebox.showinfo("完成", "Excel文件处理完成！")
            else:
                self.update_status("处理被取消")

        except Exception as e:
            self.update_status(f"错误: {str(e)}")
            messagebox.showerror("错误", f"处理过程中出现错误：\n{str(e)}")
        finally:
            self.reset_ui()

    def start_processing(self):
        if not self.file_path.get():
            messagebox.showwarning("警告", "请先选择Excel文件！")
            return

        self.is_running = True
        self.start_btn["state"] = "disabled"
        self.cancel_btn["state"] = "normal"
        self.browse_btn["state"] = "disabled"

        # 清空状态显示
        self.status_text.delete(1.0, "end")

        # 在新线程中运行处理过程
        thread = threading.Thread(target=self.format_excel)
        thread.start()

    def cancel_processing(self):
        self.is_running = False
        self.cancel_btn["state"] = "disabled"

    def reset_ui(self):
        self.start_btn["state"] = "normal"
        self.cancel_btn["state"] = "disabled"
        self.browse_btn["state"] = "normal"
        self.total_progress["value"] = 0
        self.sheet_progress["value"] = 0
        self.time_label["text"] = "预计剩余时间: --:--"


def main():
    root = tk.Tk()
    app = ExcelFormatterGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
