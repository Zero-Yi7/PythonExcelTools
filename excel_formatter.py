from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from tqdm import tqdm
import time
from datetime import datetime, timedelta
import colorama
from colorama import Fore, Style

# 初始化colorama
colorama.init()


def print_step(message, color=Fore.CYAN):
    """打印带颜色的步骤信息"""
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"{color}[{timestamp}] {message}{Style.RESET_ALL}")


def format_excel(file_path):
    try:
        print_step("开始加载Excel文件...", Fore.GREEN)
        wb = load_workbook(file_path)
        total_sheets = len(wb.sheetnames)

        # 创建总进度条
        with tqdm(
            total=total_sheets,
            desc="总体进度",
            bar_format="{l_bar}%s{bar}%s{r_bar}" % (Fore.GREEN, Style.RESET_ALL),
        ) as total_pbar:
            start_time = time.time()

            # 遍历所有工作表
            for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
                print_step(f"\n开始处理工作表: {sheet_name}", Fore.YELLOW)
                ws = wb[sheet_name]

                # 创建工作表处理步骤进度条
                steps = ["设置首行格式", "设置对齐方式", "添加框线", "调整行高列宽"]

                with tqdm(
                    total=len(steps),
                    desc="工作表进度",
                    bar_format="{l_bar}%s{bar}%s{r_bar}" % (Fore.BLUE, Style.RESET_ALL),
                ) as sheet_pbar:

                    # 1. 冻结首行和筛选
                    print_step("→ 设置冻结和筛选", Fore.CYAN)
                    ws.freeze_panes = "A2"
                    ws.auto_filter.ref = ws.dimensions

                    # 2. 设置首行格式
                    print_step("→ 设置首行格式", Fore.CYAN)
                    first_row_font = Font(bold=True, color="000000")
                    first_row_fill = PatternFill(
                        start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
                    )

                    for cell in ws[1]:
                        cell.font = first_row_font
                        cell.fill = first_row_fill
                    sheet_pbar.update(1)

                    # 3. 设置边框样式
                    thin_border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"),
                    )

                    # 获取已使用范围
                    max_row = ws.max_row
                    max_col = ws.max_column

                    # 4. 设置对齐方式
                    print_step("→ 设置对齐方式", Fore.CYAN)
                    for row in ws.iter_rows(
                        min_row=1, max_row=max_row, max_col=max_col
                    ):
                        for cell in row:
                            cell.alignment = Alignment(
                                horizontal="center", vertical="center"
                            )
                    sheet_pbar.update(1)

                    # 5. 添加框线
                    print_step("→ 添加框线", Fore.CYAN)
                    for row in ws.iter_rows(
                        min_row=1, max_row=max_row, max_col=max_col
                    ):
                        for cell in row:
                            cell.border = thin_border
                    sheet_pbar.update(1)

                    # 6. 调整行高和列宽
                    print_step("→ 调整行高和列宽", Fore.CYAN)
                    # 调整列宽
                    for col in range(1, max_col + 1):
                        max_length = 0
                        column = get_column_letter(col)
                        for row in range(1, max_row + 1):
                            cell = ws[f"{column}{row}"]
                            if cell.value:
                                cell_length = sum(
                                    2 if ord(char) > 127 else 1
                                    for char in str(cell.value)
                                )
                                max_length = max(max_length, cell_length)
                        ws.column_dimensions[column].width = max_length + 2

                    # 调整行高
                    for row in range(1, max_row + 1):
                        ws.row_dimensions[row].height = 20
                    sheet_pbar.update(1)

                # 更新总进度
                total_pbar.update(1)

                # 计算并显示预计剩余时间
                elapsed_time = time.time() - start_time
                sheets_remaining = total_sheets - sheet_idx
                if sheet_idx > 0:
                    time_per_sheet = elapsed_time / sheet_idx
                    eta = time_per_sheet * sheets_remaining
                    eta_str = str(timedelta(seconds=int(eta)))
                    print_step(f"预计剩余时间: {eta_str}", Fore.MAGENTA)

                # 显示完成百分比
                percentage = (sheet_idx / total_sheets) * 100
                print_step(f"总体完成度: {percentage:.1f}%", Fore.GREEN)

        # 保存文件
        print_step("\n正在保存文件...", Fore.YELLOW)
        wb.save(file_path)
        print_step("Excel文件处理完成！", Fore.GREEN)

    except Exception as e:
        print_step(f"处理出错: {str(e)}", Fore.RED)

    finally:
        input("\n按回车键退出...")


if __name__ == "__main__":
    try:
        file_path = input("请输入Excel文件路径: ")
        format_excel(file_path)
    except KeyboardInterrupt:
        print_step("\n处理被用户中断", Fore.YELLOW)
    except Exception as e:
        print_step(f"发生错误: {str(e)}", Fore.RED)
